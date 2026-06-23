from datetime import datetime, timezone
from io import BytesIO

import pytest
from openpyxl import load_workbook

from app.application.executions.export import (
    InstanceExportContext,
    build_all_instances_excel_export,
    build_instance_excel_export,
    safe_export_filename,
)
from app.domain.enums import ExecutionStatus, NodeStatus, WorkflowStatus


class _Instance:
    id = "inst-1"
    name = "Demo Run"
    workflow_definition_id = "wf-def-1"
    status = WorkflowStatus.COMPLETED
    created_at = datetime(2026, 6, 22, 10, 0, tzinfo=timezone.utc)
    completed_at = datetime(2026, 6, 22, 11, 0, tzinfo=timezone.utc)
    current_revision = 3
    created_by = "admin"


class _NodeInstance:
    def __init__(self, workflow_node_id: str, status: NodeStatus, current_execution: int):
        self.id = f"node-inst-{workflow_node_id}"
        self.workflow_node_id = workflow_node_id
        self.node_definition_version_id = "ver-1"
        self.status = status
        self.current_execution = current_execution
        self.created_at = datetime(2026, 6, 22, 10, 5, tzinfo=timezone.utc)
        self.updated_at = datetime(2026, 6, 22, 10, 30, tzinfo=timezone.utc)


class _Execution:
    def __init__(self, workflow_node_instance_id: str, workflow_node_id: str):
        self.workflow_node_instance_id = workflow_node_instance_id
        self.workflow_instance_id = "inst-1"
        self.execution_number = 1
        self.inputs_json = {"qty": 10}
        self.outputs_json = {"lineTotal": 15.5, "qty": 10}
        self.status = ExecutionStatus.COMPLETED
        self.executed_by = "admin"
        self.started_at = datetime(2026, 6, 22, 10, 20, tzinfo=timezone.utc)
        self.completed_at = datetime(2026, 6, 22, 10, 25, tzinfo=timezone.utc)
        self._workflow_node_id = workflow_node_id


class _Event:
    sequence_number = 1
    event_type = "WORKFLOW_STARTED"
    created_at = datetime(2026, 6, 22, 10, 0, tzinfo=timezone.utc)
    payload_json = {"workflow_definition_id": "wf-def-1"}
    created_by = "admin"


def test_safe_export_filename_sanitizes_invalid_characters():
    assert safe_export_filename("Run #1 / test!") == "Run 1  test"


def test_build_instance_excel_export_creates_expected_sheets():
    node = _NodeInstance("task-1", NodeStatus.COMPLETED, 1)
    execution = _Execution(node.id, node.workflow_node_id)
    state = {
        "task_names": {"task-1": "Raw Material Pricing"},
        "node_instances": [node],
        "execution_summary": {
            "total": 15.5,
            "items": [
                {
                    "task_name": "Raw Material Pricing",
                    "workflow_node_id": "task-1",
                    "output_label": "Line total",
                    "output_key": "lineTotal",
                    "value": 15.5,
                }
            ],
        },
        "workflow_projection": {
            "nodes": {
                "task-1": {
                    "status": "COMPLETED",
                    "cost_contribution": 15.5,
                    "outputs": {"lineTotal": 15.5},
                }
            },
            "total": 15.5,
        },
    }

    content = build_instance_excel_export(
        instance=_Instance(),
        state=state,
        executions=[execution],
        events=[_Event()],
        workflow_definition_name="Test Workflow",
    )

    assert content[:2] == b"PK"
    workbook = load_workbook(BytesIO(content))
    assert workbook.sheetnames == [
        "Summary",
        "Tasks",
        "Outputs",
        "Cost Breakdown",
        "Events",
    ]
    assert workbook["Summary"]["B7"].value == 15.5
    assert workbook["Tasks"]["A2"].value == "Raw Material Pricing"
    assert workbook["Outputs"]["A2"].value == "Raw Material Pricing"
    assert workbook["Cost Breakdown"]["A2"].value == "Raw Material Pricing"


def test_build_all_instances_excel_export_creates_expected_sheets():
    node = _NodeInstance("task-1", NodeStatus.COMPLETED, 1)
    execution = _Execution(node.id, node.workflow_node_id)
    state = {
        "task_names": {"task-1": "Raw Material Pricing"},
        "node_instances": [node],
        "execution_summary": {
            "total": 15.5,
            "items": [
                {
                    "task_name": "Raw Material Pricing",
                    "workflow_node_id": "task-1",
                    "output_label": "Line total",
                    "output_key": "lineTotal",
                    "value": 15.5,
                }
            ],
        },
        "workflow_projection": {
            "nodes": {
                "task-1": {
                    "status": "COMPLETED",
                    "cost_contribution": 15.5,
                }
            },
            "total": 15.5,
        },
    }
    context = InstanceExportContext(
        instance=_Instance(),
        state=state,
        executions=[execution],
        events=[_Event()],
        workflow_definition_name="Test Workflow",
    )

    content = build_all_instances_excel_export([context])
    workbook = load_workbook(BytesIO(content))
    assert workbook.sheetnames == [
        "Instances",
        "Tasks",
        "Outputs",
        "Cost Breakdown",
        "Events",
    ]
    assert workbook["Instances"]["A2"].value == "inst-1"
    assert workbook["Instances"]["B2"].value == "Demo Run"
    assert workbook["Tasks"]["C2"].value == "Raw Material Pricing"
