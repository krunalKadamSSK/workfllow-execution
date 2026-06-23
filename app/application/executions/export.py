from __future__ import annotations

import json
import re
from dataclasses import dataclass
from datetime import datetime
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.infrastructure.db.models.events import WorkflowEvent
from app.infrastructure.db.models.instances import WorkflowInstance, WorkflowNodeExecution


ALL_EXPORT_FILENAME = "workflow-executions-export.xlsx"


def safe_export_filename(instance_name: str) -> str:
    cleaned = re.sub(r"[^\w\- ]+", "", instance_name).strip()
    return (cleaned or "workflow-export")[:80]


def _cell_value(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, (str, int, float, bool)):
        return value
    if isinstance(value, datetime):
        return value.isoformat()
    return json.dumps(value, default=str)


def _write_sheet(
    worksheet: Worksheet,
    headers: list[str],
    rows: list[list[Any]],
) -> None:
    worksheet.append(headers)
    for row in rows:
        worksheet.append([_cell_value(value) for value in row])

    for column_index, header in enumerate(headers, start=1):
        column = get_column_letter(column_index)
        max_length = len(header)
        for row in worksheet.iter_rows(
            min_row=2,
            max_row=worksheet.max_row,
            min_col=column_index,
            max_col=column_index,
        ):
            cell_value = row[0].value
            if cell_value is not None:
                max_length = max(max_length, len(str(cell_value)))
        worksheet.column_dimensions[column].width = min(max_length + 2, 60)


def _new_workbook() -> Workbook:
    workbook = Workbook()
    default_sheet = workbook.active
    if default_sheet is not None:
        workbook.remove(default_sheet)
    return workbook


def _save_workbook(workbook: Workbook) -> bytes:
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


@dataclass(frozen=True)
class InstanceExportContext:
    instance: WorkflowInstance
    state: dict[str, Any]
    executions: list[WorkflowNodeExecution]
    events: list[WorkflowEvent]
    workflow_definition_name: str | None


def _instance_prefix(
    context: InstanceExportContext,
    *,
    include_in_all_export: bool,
) -> list[Any]:
    if not include_in_all_export:
        return []
    return [context.instance.id, context.instance.name]


def _build_instance_rows(context: InstanceExportContext) -> list[list[Any]]:
    execution_summary: dict[str, Any] = context.state.get("execution_summary") or {}
    return [
        [
            context.instance.id,
            context.instance.name,
            context.workflow_definition_name or context.instance.workflow_definition_id,
            context.instance.status.value,
            context.instance.created_at,
            context.instance.completed_at or "",
            execution_summary.get("total"),
            context.instance.current_revision,
            context.instance.created_by or "",
        ]
    ]


def _build_task_rows(
    context: InstanceExportContext,
    *,
    include_in_all_export: bool,
) -> list[list[Any]]:
    task_names: dict[str, str] = context.state.get("task_names") or {}
    projection_nodes: dict[str, Any] = (context.state.get("workflow_projection") or {}).get(
        "nodes", {}
    )
    rows: list[list[Any]] = []
    for node in context.state.get("node_instances") or []:
        node_projection = projection_nodes.get(node.workflow_node_id, {})
        rows.append(
            [
                *_instance_prefix(context, include_in_all_export=include_in_all_export),
                task_names.get(node.workflow_node_id, node.workflow_node_id),
                node.workflow_node_id,
                node.status.value,
                node.current_execution,
                node_projection.get("cost_contribution"),
                node.created_at,
                node.updated_at,
            ]
        )
    return rows


def _build_output_rows(
    context: InstanceExportContext,
    *,
    include_in_all_export: bool,
) -> list[list[Any]]:
    task_names: dict[str, str] = context.state.get("task_names") or {}
    node_instance_by_id = {
        node.id: node for node in (context.state.get("node_instances") or [])
    }
    rows: list[list[Any]] = []
    for execution in context.executions:
        node_instance = node_instance_by_id.get(execution.workflow_node_instance_id)
        if node_instance is None:
            continue
        task_name = task_names.get(
            node_instance.workflow_node_id,
            node_instance.workflow_node_id,
        )
        for field_name, field_value in execution.outputs_json.items():
            rows.append(
                [
                    *_instance_prefix(context, include_in_all_export=include_in_all_export),
                    task_name,
                    node_instance.workflow_node_id,
                    execution.execution_number,
                    field_name,
                    field_value,
                    execution.executed_by or "",
                    execution.completed_at or execution.started_at,
                ]
            )
    return rows


def _build_cost_rows(
    context: InstanceExportContext,
    *,
    include_in_all_export: bool,
) -> list[list[Any]]:
    execution_summary: dict[str, Any] = context.state.get("execution_summary") or {}
    rows: list[list[Any]] = []
    for item in execution_summary.get("items") or []:
        rows.append(
            [
                *_instance_prefix(context, include_in_all_export=include_in_all_export),
                item.get("task_name") or item.get("task_label"),
                item.get("workflow_node_id"),
                item.get("output_label"),
                item.get("output_key"),
                item.get("value"),
            ]
        )
    if not include_in_all_export:
        rows.append(["", "", "", "Total", execution_summary.get("total")])
    return rows


def _build_event_rows(
    context: InstanceExportContext,
    *,
    include_in_all_export: bool,
) -> list[list[Any]]:
    return [
        [
            *_instance_prefix(context, include_in_all_export=include_in_all_export),
            event.sequence_number,
            event.event_type,
            event.created_at,
            json.dumps(event.payload_json, default=str),
            event.created_by or "",
        ]
        for event in context.events
    ]


def build_all_instances_excel_export(contexts: list[InstanceExportContext]) -> bytes:
    """Build a multi-sheet Excel workbook exporting all workflow instances."""
    workbook = _new_workbook()

    instance_rows: list[list[Any]] = []
    task_rows: list[list[Any]] = []
    output_rows: list[list[Any]] = []
    cost_rows: list[list[Any]] = []
    event_rows: list[list[Any]] = []

    for context in contexts:
        instance_rows.extend(_build_instance_rows(context))
        task_rows.extend(_build_task_rows(context, include_in_all_export=True))
        output_rows.extend(_build_output_rows(context, include_in_all_export=True))
        cost_rows.extend(_build_cost_rows(context, include_in_all_export=True))
        event_rows.extend(_build_event_rows(context, include_in_all_export=True))

    _write_sheet(
        workbook.create_sheet("Instances"),
        [
            "Instance ID",
            "Instance Name",
            "Workflow Definition",
            "Status",
            "Created At",
            "Completed At",
            "Total Cost",
            "Revision",
            "Created By",
        ],
        instance_rows,
    )
    _write_sheet(
        workbook.create_sheet("Tasks"),
        [
            "Instance ID",
            "Instance Name",
            "Task Name",
            "Workflow Node ID",
            "Status",
            "Execution Count",
            "Cost Contribution",
            "Created At",
            "Updated At",
        ],
        task_rows,
    )
    _write_sheet(
        workbook.create_sheet("Outputs"),
        [
            "Instance ID",
            "Instance Name",
            "Task Name",
            "Workflow Node ID",
            "Execution #",
            "Field",
            "Value",
            "Executed By",
            "Completed At",
        ],
        output_rows,
    )
    _write_sheet(
        workbook.create_sheet("Cost Breakdown"),
        [
            "Instance ID",
            "Instance Name",
            "Task Name",
            "Workflow Node ID",
            "Output Label",
            "Output Key",
            "Value",
        ],
        cost_rows,
    )
    _write_sheet(
        workbook.create_sheet("Events"),
        [
            "Instance ID",
            "Instance Name",
            "Sequence",
            "Event Type",
            "Created At",
            "Payload",
            "Created By",
        ],
        event_rows,
    )

    return _save_workbook(workbook)


def build_instance_excel_export(
    *,
    instance: WorkflowInstance,
    state: dict[str, Any],
    executions: list[WorkflowNodeExecution],
    events: list[WorkflowEvent],
    workflow_definition_name: str | None,
) -> bytes:
    """Build a multi-sheet Excel workbook for a single workflow instance export."""
    context = InstanceExportContext(
        instance=instance,
        state=state,
        executions=executions,
        events=events,
        workflow_definition_name=workflow_definition_name,
    )
    workbook = _new_workbook()

    summary_sheet = workbook.create_sheet("Summary")
    summary_rows = [
        ["Instance ID", context.instance.id],
        ["Instance Name", context.instance.name],
        ["Workflow Definition", context.workflow_definition_name or context.instance.workflow_definition_id],
        ["Status", context.instance.status.value],
        ["Created At", context.instance.created_at],
        ["Completed At", context.instance.completed_at or ""],
        ["Total Cost", (context.state.get("execution_summary") or {}).get("total")],
        ["Revision", context.instance.current_revision],
        ["Created By", context.instance.created_by or ""],
    ]
    for row_index, (label, value) in enumerate(summary_rows, start=1):
        summary_sheet.cell(row=row_index, column=1, value=label)
        summary_sheet.cell(row=row_index, column=2, value=_cell_value(value))
    summary_sheet.column_dimensions["A"].width = 24
    summary_sheet.column_dimensions["B"].width = 48

    _write_sheet(
        workbook.create_sheet("Tasks"),
        [
            "Task Name",
            "Workflow Node ID",
            "Status",
            "Execution Count",
            "Cost Contribution",
            "Created At",
            "Updated At",
        ],
        _build_task_rows(context, include_in_all_export=False),
    )
    _write_sheet(
        workbook.create_sheet("Outputs"),
        [
            "Task Name",
            "Workflow Node ID",
            "Execution #",
            "Field",
            "Value",
            "Executed By",
            "Completed At",
        ],
        _build_output_rows(context, include_in_all_export=False),
    )

    execution_summary: dict[str, Any] = context.state.get("execution_summary") or {}
    cost_rows = _build_cost_rows(context, include_in_all_export=False)
    cost_rows.append(["", "", "", "Total", execution_summary.get("total")])
    _write_sheet(
        workbook.create_sheet("Cost Breakdown"),
        ["Task Name", "Workflow Node ID", "Output Label", "Output Key", "Value"],
        cost_rows,
    )
    _write_sheet(
        workbook.create_sheet("Events"),
        ["Sequence", "Event Type", "Created At", "Payload", "Created By"],
        _build_event_rows(context, include_in_all_export=False),
    )

    return _save_workbook(workbook)
